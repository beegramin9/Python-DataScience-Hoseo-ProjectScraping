import datetime
dt = datetime.datetime.now()
df = dt.strftime('%Y_%m_%d')

import requests
from bs4 import BeautifulSoup

url = 'https://hashcode.co.kr/?page='

tag_title = {}

for page in range(1):
    res = requests.get(url+str(page+1))
    raw = res.text
    html = BeautifulSoup(raw,'html.parser')

    questions = html.select('li.question-list-item')
    
    for question in questions:
        tags = question.select('li.label.label-tag a > span')
        '''
        이상하게 tag = question.select_one().text는
        span까지 분석하여 text로 양옆 태그만 날려주면 되는 상황임에도 불구하고
        .text가 안 먹힘
        '''
        title = question.select_one('div.top a').text
        for tag in tags:
            tag = tag.text
            '''
            그래서 우선 tags / select로 리스트까지만 만들어 놓고
            반복문을 사용해서 tags 안의 tag를 일일히 꺼내고 .text로 만들어 주니까 됨 
            '''
            # 여기까지 오면 각각 tag하나(반복문의 첫 타), title 하나(select_one의 첫 타)
            # 가 모여서 한 쌍의 tag_title[tag]를 만들 수 있음
            # 그래서 동작하는 거임
            if tag in tag_title.keys():
                tag_title[tag].append(title)
            else:
                tag_title[tag] = [title]

import openpyxl
excel = openpyxl.Workbook()
sheet = excel.active
sheet.title = 'Question list'

row = 1
for item in tag_title.items():
    sheet.cell(row = row, column = 1).value = item[0]
    for title in item[1]:
        sheet.cell(row=row, column = 2).value = title
        row += 1
    row += 1 

excel.save('10_hashcode_tag_'+df+'.xlsx')