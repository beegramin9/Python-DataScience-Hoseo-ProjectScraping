import openpyxl

wb = openpyxl.Workbook()
sheet1 = wb['Sheet']
sheet1.title = 'openpyxl 연습'

sheet1['A1'] = 'Hello openpyxl!'
days = ['월','화','수','목','금','토','일']
for i in days:
    sheet1.cell(row=2,column = days.index(i)+1).value = i +'요일'

sheet2 = wb.create_sheet('계단식 배치')

for i in range(10):
    sheet2.cell(row=i+1,column=i+1).value = i+1


wb.save('5_openpyxl_practice.xlsx')
