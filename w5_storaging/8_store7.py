import requests
from bs4 import BeautifulSoup
# 검색어도 변경 가능
search = input('검색어를 입력해 주세요: ')

url = '''
https://search.naver.com/search.naver?where=news
&sm=tab_jum
&query=
'''
# https://book.coalastudy.com/data-crawling/week-5/stage-4

"""
하나의 키의 값이 리스트로 되어있고 그 리스트에 여러개의 값이 들어있을 때
엑셀로 어떻게 저장할 수 있을까?
"""


info_dic = {}

for i in range(3):
    raw = requests.get(url+search+'&start='+str(i*10+1),
    headers = {'User-Agent':'Mozilla/5.0'}).text
    html = BeautifulSoup(raw,'html.parser')
    articles = html.select('ul.type01 > li')

    for article in articles:
        # 뽑을 게 여러개 있을 때는 전체를 아우르는 요소를 찾아야지
        # 반복문을 한 번만 돌릴 수 있다.
        press = article.select_one('span._sp_each_source').text
        title = article.select_one('a._sp_each_title').text
        titlelist = []
        if press in info_dic.keys():
            # items는 튜플이라고!!!! ㅠㅠㅠㅠ
            # itmes에 당연히 key가 없지. items는 튜플이고 key는 문자열인데
            # 그러니까 info_dic.items()로 하면 전부 else로 가서 겹쳐지지 않고 따로 만드는 거야
            info_dic[press].append(title)
        else:
            # info_dic = {press:titlelist} 이건 도대체 어디서 나온 문법이냐...
            info_dic[press] = [title]
            '''
            dic[key] = value 하면 value값 하나만 나오지만
            dic[key] = [value] 하면 value를 감싼 리스트로 나오는 거 처음 앎
            '''
# for item in info_dic.items():
#     print(item)

import openpyxl
excel = openpyxl.Workbook()
sheet = excel.active
row = 1

for item in info_dic.items():
    sheet.cell(row = row, column = 1).value = item[0]
    for title in item[1]:
        sheet.cell(row = row, column = 2).value = title
        row += 1 
    row += 1 

excel.save('8_store7_'+search+'.xlsx')