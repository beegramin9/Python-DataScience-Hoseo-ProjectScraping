import openpyxl

excel = openpyxl.Workbook()
clips_sheet = excel.active
clips_sheet.title = 'Top 100 Clips'
clips_sheet.append(['Title', 'Channel', 'View', 'Like'])

by_hit_sheet = excel.create_sheet('By_Hit')
by_hit_sheet.append(['Channel', 'Total Hit', 'Total Like'])

by_like_sheet = excel.create_sheet('By_Like')
by_like_sheet.append(['Channel', 'Total Like', 'Total Hit'])

import requests
from bs4 import BeautifulSoup

res = requests.get('https://tv.naver.com/r')
raw = res.text
html = BeautifulSoup(raw, 'html.parser')

clips = html.select('dl.cds_info')
# int 반드시!
'''
딕셔너리에서 숫자를 계속 더하니까 반드시 int 해줘야 함
'''
# Top 100 Clips
for clip in clips:
    title = clip.select_one('dt.title a').text.strip()
    chn = clip.select_one('dd.chn').text
    hit = int(clip.select_one('span.hit').text.replace(',', '')[4:])
    like = int(clip.select_one('span.like').text.replace(',', '')[5:])

    clips_sheet.append([title, chn, hit, like])

# By_Hit,By_Like
clips_dic = {}
for clip in clips:
    chn = clip.select_one('dd.chn').text.strip()
    hit = int(clip.select_one('span.hit').text.replace(',', '')[4:])
    like = int(clip.select_one('span.like').text.replace(',', '')[5:])

    if chn in clips_dic.keys():
        clips_dic[chn]['hit'] += hit
        clips_dic[chn]['like'] += like
    else:
        clips_dic[chn] = {'hit': hit, 'like': like}

# 순위변경
# 1-3위
up_down_list = []
delta_list = []

top3s = html.select('div.info')

for top3 in top3s:
    change = top3.select_one('span.blind').text
    up_down_list.append(change)

    delta = top3.select_one('em').text
    delta_list.append(delta)

from4s = html.select('div.cds_type')

for from4 in from4s:
    delta = from4.select_one('em').text
    delta_list.append(delta)
    print(delta)
    change = from4.select('span.blind')[1].text
    # 재생 수 라고 나와있으면 new(신규진입)임
    change = '신규' if change == '재생 수' else change
    up_down_list.append(change)

import pandas as pd
up_down_change = pd.DataFrame({'Change': up_down_list, 'Delta': delta_list})
up_down_change.to_csv('change_column.csv', sep=',', encoding='utf8')


def sort_by_hit(item):
    return item[1]['hit']


def sort_by_like(item):
    return item[1]['like']


sorted_by_hit = sorted(clips_dic.items(), key=sort_by_hit, reverse=True)
sorted_by_like = sorted(clips_dic.items(), key=sort_by_like, reverse=True)

for sortedclip in sorted_by_hit:
    by_hit_sheet.append(
        [sortedclip[0], sortedclip[1]['hit'], sortedclip[1]['like']])

for sortedclip in sorted_by_like:
    by_like_sheet.append(
        [sortedclip[0], sortedclip[1]['like'], sortedclip[1]['hit']])

import datetime
dt = datetime.datetime.now()
df = dt.strftime('%Y_%m_%d')

filename = '6_naver_tv_100_' + df

excel.save(filename + '.xlsx')
excel.close()