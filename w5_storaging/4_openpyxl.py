'''
csv파일은 엑셀의 수식,서식,필터,차트 등의 기능을 사용하지 못한다
'''

import openpyxl
'''
Workbook = 엑셀 파일 전체
Worksheet = 엑셀의 시트 하나하나
Cell = 엑셀의 칸 하나하나
'''
# openpyxl 타입함수
wb1 = openpyxl.Workbook()

sheet = wb1.active # .active는 활성화된 시트를 불러옴
# 타입함수일 텐데 왜 ()가 안 붙는지는 의문
# 지금까지는 시트가 하나뿐이기 때문에 첫 번째 시트 생성
sheet2 = wb1.create_sheet('sheet2')
# 새로운 Sheet 만들기

sheet2_updated = wb1['sheet2']
# 이미 존재하면 딕셔너리처럼 불러올 수 있음
# 내가 임의로 sheet2_updated를 불러온 것

sheet2_updated.title = '수집 데이터'
# Sheet 이름 바꾸기

wb1.save('4_openpyxl1.xlsx') # f.close()와 같은 역할

wb2 = openpyxl.Workbook()
sheet = wb2.active

sheet['B2'] = 'b2' # 열-행
sheet.cell(row=3,column=3).value = '3,3'
'''
굳이 이렇게 쓰는 이유는? 반복문 쓰기에 좋다
'''
sheet['A2'] = '검정고무신'
sheet['B2'] = '검정고무신'
sheet['C2'] = '검정고무신'
sheet['D2'] = '검정고무신'
sheet['E2'] = '검정고무신'

for i in range(5):
    sheet.cell(row=1,column = i+1).value = '검정고무신'

sheet.append([1,2,3,4,5])

wb2.save('4_openpyxl2.xlsx')

wb3 = openpyxl.Workbook()
sheet1 = wb3['Sheet']
# 기본으로 설정되는 첫 번째 이름이 sheet. 즉 이미 존재하기 떄문에 딕셔너리처럼 불러올 수 있다.
sheet1.title = '수집 데이터'
sheet1['A1'] = '첫 번째 시트'

sheet2 = wb3.create_sheet('정리 결과')
sheet2.cell(row=1, column=1).value = '두 번째 시트'

sheet1.append(['다시','첫 번째 시트'])
wb3.save('4_openpyxl3.xlsx')

wb4 = openpyxl.load_workbook('4_openpyxl1.xlsx')
sheet1 = wb4.active
sheet1.title = '이름 변경'
sheet1.append(range(10))

wb4.save('4_openpyxl4.xlsx')